from __future__ import annotations

import csv
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from getpass import getpass
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from playwright.sync_api import Browser, BrowserContext, Page, Playwright, sync_playwright

DOC24_HOME = "https://docu.gdoc.go.kr/index.do"
SENT_DOCS_URL = "https://docu.gdoc.go.kr/doc/snd/sendDocList.do"

APP_DIR = Path.home() / ".doc24_sender"
CONFIG_PATH = APP_DIR / "config.json"
RESULT_DIR = APP_DIR / "results"

CHROME_PATHS = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    str(Path.home() / "Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
]


@dataclass
class RecipientResult:
    recipient: str
    status: str
    reason: str = ""


def log(message: str) -> None:
    print(f"[{datetime.now():%H:%M:%S}] {message}", flush=True)


def ensure_app_dirs() -> None:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    RESULT_DIR.mkdir(parents=True, exist_ok=True)


def parse_recipients(text: str) -> list[str]:
    recipients: list[str] = []
    seen: set[str] = set()
    for raw in text.replace(",", "\n").splitlines():
        value = raw.strip()
        if value and value not in seen:
            recipients.append(value)
            seen.add(value)
    return recipients


def load_recipients() -> list[str]:
    candidates = [
        Path("school_list.xlsx"),
        Path("recipients.xlsx"),
        Path("recipients.csv"),
        Path("recipients.txt"),
    ]
    source = next((path for path in candidates if path.exists()), None)
    if source is None:
        raise FileNotFoundError(
            "수신기관 파일이 없습니다. school_list.xlsx, recipients.xlsx, recipients.csv, recipients.txt 중 하나를 넣어주세요."
        )

    values: list[str] = []
    suffix = source.suffix.lower()

    if suffix == ".txt":
        values = source.read_text(encoding="utf-8-sig").splitlines()
    elif suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                if row and row[0] is not None:
                    values.append(str(row[0]).strip())
    else:
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] is not None:
                values.append(str(row[0]).strip())
        workbook.close()

    names = parse_recipients("\n".join(values))
    if names and names[0].replace(" ", "") in {"학교명", "기관명", "수신기관", "수신자"}:
        names = names[1:]

    if not names:
        raise RuntimeError(f"{source.name}에서 수신기관을 찾지 못했습니다.")

    log(f"수신기관 {len(names)}개 로드: {source.name}")
    return names


def save_results(results: list[RecipientResult]) -> Path:
    ensure_app_dirs()
    output = RESULT_DIR / f"문서24_발송결과_{datetime.now():%Y%m%d_%H%M%S}.csv"
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["수신기관", "상태", "사유"])
        for item in results:
            writer.writerow([item.recipient, item.status, item.reason])
    return output


def load_local_credentials() -> tuple[str, str] | None:
    if not CONFIG_PATH.exists():
        return None
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        username = str(data.get("username", "")).strip()
        password = str(data.get("password", ""))
        if username and password:
            return username, password
    except Exception:
        return None
    return None


def save_local_credentials(username: str, password: str) -> None:
    ensure_app_dirs()
    CONFIG_PATH.write_text(
        json.dumps({"username": username, "password": password}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    try:
        os.chmod(CONFIG_PATH, 0o600)
    except OSError:
        pass


def extract_legacy_credentials(source: str) -> tuple[str, str] | None:
    id_patterns = [
        r'page\.fill\(\s*["\']#id["\']\s*,\s*["\']([^"\']+)["\']\s*\)',
        r'locator\(\s*["\']#id["\']\s*\)\.fill\(\s*["\']([^"\']+)["\']\s*\)',
    ]
    pw_patterns = [
        r'page\.keyboard\.type\(\s*["\']([^"\']+)["\']',
        r'locator\(\s*["\']#password["\']\s*\)\.fill\(\s*["\']([^"\']+)["\']\s*\)',
        r'page\.fill\(\s*["\']#password["\']\s*,\s*["\']([^"\']+)["\']\s*\)',
    ]

    username = next(
        (match.group(1) for pattern in id_patterns if (match := re.search(pattern, source))),
        None,
    )
    password = next(
        (match.group(1) for pattern in pw_patterns if (match := re.search(pattern, source))),
        None,
    )
    if username and password:
        return username, password
    return None


def migrate_legacy_credentials() -> tuple[str, str] | None:
    sources: list[str] = []

    candidates = [
        Path.home() / "Desktop" / "doc24-main-backup.py",
        Path("legacy_main.py"),
        Path("main_old.py"),
        Path("old_main.py"),
    ]
    for path in candidates:
        if path.exists():
            try:
                sources.append(path.read_text(encoding="utf-8"))
            except Exception:
                pass

    try:
        completed = subprocess.run(
            ["git", "show", "main:main.py"],
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        )
        if completed.stdout:
            sources.append(completed.stdout)
    except Exception:
        pass

    for source in sources:
        credentials = extract_legacy_credentials(source)
        if credentials:
            save_local_credentials(*credentials)
            log("기존 코드의 로그인 정보를 맥 로컬 설정으로 저장했습니다.")
            return credentials
    return None


def get_credentials() -> tuple[str, str]:
    credentials = load_local_credentials() or migrate_legacy_credentials()
    if credentials:
        return credentials

    print("\n문서24 로그인 정보가 아직 로컬에 저장되어 있지 않습니다.")
    username = input("아이디: ").strip()
    password = getpass("비밀번호: ")
    if not username or not password:
        raise RuntimeError("아이디 또는 비밀번호가 비어 있습니다.")
    save_local_credentials(username, password)
    print("로그인 정보를 이 맥의 ~/.doc24_sender/config.json 에 저장했습니다.\n")
    return username, password


def normalize_org_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", value).lower()


class PreventSleep:
    def __init__(self) -> None:
        self.process: subprocess.Popen | None = None

    def __enter__(self):
        try:
            self.process = subprocess.Popen(
                ["caffeinate", "-dimsu"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception:
            self.process = None
        return self

    def __exit__(self, exc_type, exc, tb):
        if self.process is not None:
            self.process.terminate()
            try:
                self.process.wait(timeout=2)
            except Exception:
                self.process.kill()


class Doc24Automation:
    def __init__(self, logger: Callable[[str], None] = log):
        self.log = logger
        self.playwright: Playwright | None = None
        self.browser: Browser | None = None
        self.context: BrowserContext | None = None
        self.page: Page | None = None

    def __enter__(self):
        ensure_app_dirs()
        self.playwright = sync_playwright().start()
        chrome_path = next((path for path in CHROME_PATHS if Path(path).exists()), None)
        if not chrome_path:
            raise RuntimeError("Google Chrome을 찾지 못했습니다. /Applications에 Chrome을 설치해주세요.")

        self.browser = self.playwright.chromium.launch(
            executable_path=chrome_path,
            headless=False,
        )
        self.context = self.browser.new_context(viewport={"width": 1280, "height": 900})
        self.page = self.context.new_page()
        return self

    def __exit__(self, exc_type, exc, tb):
        if exc_type is not None and sys.stdin.isatty():
            print("\n오류가 발생했습니다. 현재 Chrome 화면을 확인하세요.")
            try:
                input("확인 후 Enter를 누르면 Chrome을 닫습니다: ")
            except Exception:
                pass
        try:
            if self.context is not None:
                self.context.close()
            if self.browser is not None:
                self.browser.close()
        finally:
            if self.playwright is not None:
                self.playwright.stop()

    def _login_successful(self) -> bool:
        assert self.page is not None
        try:
            return "로그아웃" in self.page.content()
        except Exception:
            return False

    def ensure_login(self) -> None:
        assert self.page is not None
        page = self.page
        username, password = get_credentials()

        self.log("문서24 로그인 페이지 이동")
        page.goto(DOC24_HOME, wait_until="domcontentloaded", timeout=30000)

        self.log("로그인 메뉴 선택")
        page.get_by_text("로그인", exact=True).click()
        page.wait_for_timeout(1200)

        self.log("법인 계정 선택")
        page.locator("#entrprsHref").click()
        page.wait_for_timeout(500)

        self.log("아이디 입력")
        page.locator("#id").fill(username)

        self.log("비밀번호 입력")
        page.locator("#password").fill(password)

        self.log("로그인 실행")
        page.locator("#password").press("Enter")

        for _ in range(20):
            page.wait_for_timeout(500)
            if self._login_successful():
                self.log("로그인 성공")
                return

        raise RuntimeError("문서24 로그인에 실패했습니다. 아이디/비밀번호 또는 로그인 화면을 확인해주세요.")

    def _save_debug(self, reason: str) -> Path:
        assert self.page is not None
        ensure_app_dirs()
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = RESULT_DIR / f"debug_{stamp}"
        try:
            self.page.screenshot(path=str(prefix.with_suffix(".png")), full_page=True)
        except Exception:
            pass
        try:
            prefix.with_suffix(".html").write_text(self.page.content(), encoding="utf-8")
        except Exception:
            pass
        try:
            prefix.with_suffix(".txt").write_text(reason, encoding="utf-8")
        except Exception:
            pass
        return prefix

    def _find_latest_document(self):
        assert self.page is not None
        page = self.page
        rows = page.locator("tbody tr")
        count = rows.count()
        notes: list[str] = []

        for index in range(count):
            row = rows.nth(index)
            try:
                row_text = " ".join(row.inner_text().split())
            except Exception:
                row_text = ""

            if not row_text:
                continue
            notes.append(f"row {index}: {row_text[:300]}")

            if "조회된 데이터가 없습니다" in row_text or "검색 결과가 없습니다" in row_text:
                continue

            anchors = row.locator("a")
            for anchor_index in range(anchors.count()):
                anchor = anchors.nth(anchor_index)
                try:
                    text = " ".join(anchor.inner_text().split())
                except Exception:
                    text = ""
                if text:
                    return row, anchor, text

            if anchors.count() > 0:
                return row, anchors.first, row_text

            clickable = row.locator("[onclick]")
            if clickable.count() > 0:
                target = clickable.first
                try:
                    text = " ".join(target.inner_text().split())
                except Exception:
                    text = row_text
                return row, target, text or row_text

            try:
                if row.get_attribute("onclick"):
                    return row, row, row_text
            except Exception:
                pass

        reason = "\n".join(notes[:15]) or "보낸 문서함 행을 읽지 못했습니다."
        debug = self._save_debug(reason)
        raise RuntimeError(f"마지막 전송문서를 찾지 못했습니다. 디버그: {debug}")

    def get_last_document_title(self) -> str:
        assert self.page is not None
        page = self.page
        self.log("보낸 문서함 이동")
        page.goto(SENT_DOCS_URL, wait_until="domcontentloaded", timeout=30000)
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        page.wait_for_timeout(1500)

        _, _, title = self._find_latest_document()
        return title

    def _click_dialog_button(self, label: str) -> None:
        assert self.page is not None
        page = self.page
        candidates = [
            page.locator(f".jconfirm-buttons button:has-text('{label}')"),
            page.locator(f"button.btnSkyBlue:has-text('{label}')"),
            page.get_by_role("button", name=label),
        ]
        for locator in candidates:
            try:
                if locator.count() and locator.last.is_visible(timeout=1200):
                    locator.last.click(force=True)
                    return
            except Exception:
                continue
        raise RuntimeError(f"'{label}' 버튼을 찾지 못했습니다.")

    def _open_last_document_for_rewrite(self) -> None:
        assert self.page is not None
        page = self.page

        page.goto(SENT_DOCS_URL, wait_until="domcontentloaded", timeout=30000)
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        page.wait_for_timeout(1500)

        _, target, _ = self._find_latest_document()
        target.click(force=True)
        page.wait_for_timeout(2500)

        rewrite = page.locator("button:has-text('재작성')")
        if rewrite.count() == 0:
            debug = self._save_debug("재작성 버튼을 찾지 못했습니다.")
            raise RuntimeError(f"재작성 버튼을 찾지 못했습니다. 디버그: {debug}")
        rewrite.first.click(force=True)
        page.wait_for_timeout(1000)

        self._click_dialog_button("예")
        page.wait_for_timeout(1800)

        for index in range(1, 5):
            checkbox = page.locator(f"label[for='wteChk{index}']")
            try:
                if checkbox.count() and checkbox.is_visible():
                    checkbox.click()
                    page.wait_for_timeout(200)
            except Exception:
                pass

        confirm = page.get_by_role("button", name="확인")
        try:
            if confirm.count() and confirm.last.is_visible(timeout=1000):
                confirm.last.click()
                page.wait_for_timeout(800)
        except Exception:
            pass

    def _select_recipient(self, recipient: str) -> None:
        assert self.page is not None
        page = self.page

        page.locator("#ldapSearch").click()
        page.wait_for_timeout(1500)

        search_box = page.locator("#searchOrgNm")
        search_box.fill(recipient)

        search_button = page.get_by_role("button", name="검색")
        if search_button.count():
            search_button.last.click()
        else:
            page.locator("button[type='submit']").last.click()
        page.wait_for_timeout(1800)

        rows = page.locator("tr:has(button:has-text('선택'))")
        row_count = rows.count()
        self.log(f"수신기관 검색 결과 {row_count}건")

        if row_count == 0:
            debug = self._save_debug(f"수신기관 검색 결과 0건: {recipient}")
            raise RuntimeError(f"수신기관 검색 결과가 없습니다. 디버그: {debug}")

        candidates: list[tuple[object, str]] = []
        for index in range(row_count):
            row = rows.nth(index)
            try:
                raw = " ".join(row.inner_text().split())
            except Exception:
                raw = ""
            if raw:
                self.log(f"검색 후보 {index + 1}: {raw}")
            candidates.append((row, raw))

        target_normalized = normalize_org_name(recipient)
        exact_matches = []
        loose_matches = []

        for row, raw in candidates:
            row_normalized = normalize_org_name(raw)
            if not row_normalized:
                continue
            if target_normalized == row_normalized:
                exact_matches.append(row)
            elif target_normalized in row_normalized:
                loose_matches.append(row)

        if len(exact_matches) == 1:
            target_row = exact_matches[0]
            self.log("수신기관 정확 일치 선택")
        elif len(loose_matches) == 1:
            target_row = loose_matches[0]
            self.log("수신기관 이름 포함 일치 선택")
        elif row_count == 1:
            target_row = candidates[0][0]
            self.log("검색 결과가 1건이라 해당 기관 선택")
        else:
            candidate_text = " | ".join(raw for _, raw in candidates if raw)
            debug = self._save_debug(
                f"수신기관 다중 결과 판단 실패\n검색어: {recipient}\n후보: {candidate_text}"
            )
            raise RuntimeError(
                "수신기관 검색 결과가 여러 개라 자동 선택하지 않았습니다. "
                f"후보를 확인해주세요. 디버그: {debug}"
            )

        target_row.get_by_role("button", name="선택").click()
        page.wait_for_timeout(700)

    def resend_to(self, recipient: str, dry_run: bool) -> RecipientResult:
        assert self.page is not None
        page = self.page
        try:
            self._open_last_document_for_rewrite()
            self._select_recipient(recipient)

            if dry_run:
                return RecipientResult(recipient, "테스트", "수신기관 검색/선택 성공")

            page.locator("#sendDoc").click()
            page.wait_for_timeout(800)
            self._click_dialog_button("보내기")
            page.wait_for_timeout(800)
            self._click_dialog_button("예")
            page.wait_for_timeout(1800)
            return RecipientResult(recipient, "완료")
        except Exception as exc:
            debug = self._save_debug(f"수신기관 처리 실패: {recipient}\n{exc}")
            if not dry_run:
                try:
                    page.keyboard.press("Escape")
                except Exception:
                    pass
            return RecipientResult(recipient, "실패", f"{exc} | 디버그: {debug}")


def run() -> int:
    ensure_app_dirs()
    recipients = load_recipients()
    dry_run = os.getenv("DOC24_DRY_RUN", "").strip().lower() in {"1", "true", "yes", "y"}

    with PreventSleep(), Doc24Automation() as automation:
        automation.ensure_login()
        last_title = automation.get_last_document_title()

        print("\n" + "=" * 70)
        print("마지막 전송문서")
        print(last_title)
        print(f"수신기관: {len(recipients)}개")
        print("모드:", "테스트(실제 발송 안 함)" if dry_run else "실제 발송")
        print("=" * 70)

        if not dry_run:
            answer = input("이 문서를 위 수신기관에 재발송할까요? [YES 입력]: ").strip()
            if answer != "YES":
                print("취소했습니다.")
                return 0
        else:
            input("테스트를 시작하려면 Enter를 누르세요: ")

        results: list[RecipientResult] = []
        total = len(recipients)

        for index, recipient in enumerate(recipients, start=1):
            log(f"[{index}/{total}] {recipient} 처리 시작")
            result = automation.resend_to(recipient, dry_run=dry_run)
            results.append(result)

            if result.status == "실패":
                log(f"실패: {recipient} - {result.reason}")
            elif result.status == "테스트":
                log(f"테스트 성공: {recipient}")
            else:
                log(f"발송 완료: {recipient}")

        output = save_results(results)
        success = sum(1 for item in results if item.status in {"완료", "테스트"})
        failed = sum(1 for item in results if item.status == "실패")

        print("\n" + "=" * 70)
        print(f"처리 완료: {success} / 실패: {failed}")
        print(f"결과 파일: {output}")
        print("=" * 70)

    return 0


def smoke_test() -> int:
    assert parse_recipients("기관A\n기관B\n기관A") == ["기관A", "기관B"]
    assert parse_recipients("기관A, 기관B") == ["기관A", "기관B"]
    assert normalize_org_name("전주 교육대학교-부설초등학교") == "전주교육대학교부설초등학교"
    sample = 'page.fill("#id", "example-user")\npage.keyboard.type("example-password", delay=100)'
    assert extract_legacy_credentials(sample) == ("example-user", "example-password")
    return 0


def main() -> int:
    if "--smoke-test" in sys.argv:
        return smoke_test()
    return run()


if __name__ == "__main__":
    raise SystemExit(main())
